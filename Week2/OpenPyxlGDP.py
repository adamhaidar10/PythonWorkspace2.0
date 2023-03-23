import openpyxl

# open a xlsx file (workbook)
workbook = openpyxl.load_workbook("gdp.xlsx")
# get the names of the sheets in the workbook
print(workbook.sheetnames)

# open a specific sheet
sheet = workbook["Data"]
# get the name of the sheet
print(sheet.title)

# get data from some cells
print(sheet["A23"].value)          # by cell name
print(sheet.cell(14, 8).value)     # by row and column numbers

# get data from a range of cells
country_names = sheet["A"]
#for country_name in country_names:
    #print(country_name.value)

country_names = sheet.iter_rows(min_col = 1, max_col = 1, min_row = 5, values_only = True)
for country_name in country_names:
    print(country_name[0])

country_names = sheet.iter_rows(min_col = 1, max_col = 64, min_row = 5, values_only = True)
for country_name in country_names:
    print("{} : {} : {}".format(country_name[0], country_name[34], country_name[63]))

# close the file when we are finished with it
workbook.close()